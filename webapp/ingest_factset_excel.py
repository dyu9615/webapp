#!/usr/bin/env python3
"""
QuantAlpha — FactSet Excel Snapshot Ingestion Script
════════════════════════════════════════════════════════
Parses FactSet-exported Excel .xlsx snapshots and ingests
structured data into bloomberg_archive.db for cross-validation
and ML Validation Gate consumption.

Supports: AAPL, DELL, JPM, MSFT, NVDA (and any FactSet format)
"""

import openpyxl
import sqlite3
import json
import os
import sys
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), 'data_service', 'bloomberg_archive.db')

# ── FactSet Excel files to process ──────────────────────────────────────────
EXCEL_FILES = {
    'AAPL': '/Users/yangyu/Downloads/webapp_Alpha_Research_Lab_v3/Snapshot_AAPL-US_2026-03-05T09_30_31_AppleInc.xlsx',
    'DELL': '/Users/yangyu/Downloads/webapp_Alpha_Research_Lab_v3/Snapshot_DELL-US_2026-03-05T09_30_09_DellTechnologiesIncC.xlsx',
    'JPM':  '/Users/yangyu/Downloads/webapp_Alpha_Research_Lab_v3/Snapshot_JPM-US_2026-03-05T09_31_25_JPMorganChase&Co.xlsx',
    'MSFT': '/Users/yangyu/Downloads/webapp_Alpha_Research_Lab_v3/Snapshot_MSFT-US_2026-03-05T09_30_56_MicrosoftCorporation.xlsx',
    'NVDA': '/Users/yangyu/Downloads/webapp_Alpha_Research_Lab_v3/Snapshot_NVDA-US_2026-03-05T09_31_07_NVIDIACorporation.xlsx',
}

COMPANY_NAMES = {
    'AAPL': 'Apple Inc.',
    'DELL': 'Dell Technologies Inc.',
    'JPM':  'JPMorgan Chase & Co.',
    'MSFT': 'Microsoft Corporation',
    'NVDA': 'NVIDIA Corporation',
}

SECTORS = {
    'AAPL': 'Technology',
    'DELL': 'Technology',
    'JPM':  'Financials',
    'MSFT': 'Technology',
    'NVDA': 'Technology',
}

INDUSTRIES = {
    'AAPL': 'Consumer Electronics',
    'DELL': 'Computer Hardware',
    'JPM':  'Diversified Banks',
    'MSFT': 'Software - Infrastructure',
    'NVDA': 'Semiconductors',
}


def parse_dollar(v):
    """Parse dollar-formatted strings like '$3,854.09' → 3854.09"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('$', '').replace(',', '').strip()
    if s in ('—', '', 'N/A'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_pct(v):
    """Parse percentage strings like '90.96' → 90.96"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('%', '').strip()
    if s in ('—', '', 'N/A'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_rating(v):
    """Parse rating string like 'Overweight (1.50)' → {'label': 'Overweight', 'score': 1.50}"""
    if v is None:
        return None, None
    s = str(v).strip()
    if '(' in s:
        label = s.split('(')[0].strip()
        try:
            score = float(s.split('(')[1].replace(')', '').strip())
        except:
            score = None
        return label, score
    return s, None


def parse_excel(ticker, filepath):
    """Parse a FactSet Excel snapshot into structured data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    kv = {}  # key-value pairs from all sections
    perf = {}  # price performance
    valuation_rows = []
    financial_rows = []
    ratio_rows = []
    comps_rows = []

    section = ''
    in_valuation = False
    in_financial = False
    in_ratio = False
    in_comps = False
    val_header = None
    fin_header = None
    ratio_header = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = list(row)
        if not cells or len(cells) < 2:
            continue

        label = cells[1]
        if label is None:
            continue
        label = str(label).strip()
        val = cells[2] if len(cells) > 2 else None

        # Price performance row (ticker line)
        if label.startswith(ticker) and len(cells) >= 7:
            perf = {
                '1M': cells[2], '3M': cells[3], '6M': cells[4],
                'YTD': cells[5], '1Y': cells[6]
            }
            continue

        # Section headers
        if label in ('Price', 'Key Stats', 'Trading', 'Current Valuation', 'Estimates',
                      'Dividends', 'Enterprise Value Bridge', 'Valuation', 'Financial Summary',
                      'Ratio', 'Revenue Exposures', 'Key Comps', 'Transactions', 'People',
                      'Profitability', 'Per Share', 'Cash Flow', 'Balance Sheet', 'Growth',
                      'Efficiency Ratios', 'Liquidity Ratios', 'Solvency Ratios'):
            section = label
            in_valuation = (label == 'Valuation')
            in_financial = (label == 'Financial Summary')
            in_ratio = (label == 'Ratio')
            if label not in ('Valuation', 'Financial Summary', 'Ratio'):
                in_valuation = in_financial = in_ratio = False
            in_comps = (label in ('Key Comps', 'Revenue Exposures'))
            continue

        # Sub-section headers within Ratio (PROFITABILITY, LIQUIDITY RATIOS, etc.)
        if in_ratio and label in ('PROFITABILITY', 'LIQUIDITY RATIOS', 'SOLVENCY RATIOS',
                                   'EFFICIENCY RATIOS', 'PER SHARE RATIOS',
                                   'INCOME STATEMENT', 'BALANCE SHEET', 'CASH FLOW'):
            continue

        # Multi-column table headers (5Yr Avg, Jan '24, etc.)
        if label == '' and val and isinstance(val, str) and '5Yr' in str(val):
            header = [str(c) if c else '' for c in cells[1:]]
            if in_valuation:
                val_header = header
            elif in_financial:
                fin_header = header
            elif in_ratio:
                ratio_header = header
            continue

        # Skip separator rows
        if label in ('', '—'):
            continue

        # Valuation table rows
        if in_valuation and val_header and label not in ('—', ''):
            row_data = {'metric': label}
            for i, h in enumerate(val_header[1:], 1):
                if h and i < len(cells) - 1:
                    row_data[h] = cells[i + 1]
            valuation_rows.append(row_data)
            continue

        # Financial summary table rows
        if in_financial and fin_header and label not in ('—', ''):
            row_data = {'metric': label}
            for i, h in enumerate(fin_header[1:], 1):
                if h and i < len(cells) - 1:
                    row_data[h] = cells[i + 1]
            financial_rows.append(row_data)
            if val is not None:
                kv[label] = val
            continue

        # Ratio table rows (profitability, liquidity, solvency, efficiency)
        if in_ratio and ratio_header and label not in ('—', ''):
            row_data = {'metric': label}
            for i, h in enumerate(ratio_header[1:], 1):
                if h and i < len(cells) - 1:
                    row_data[h] = cells[i + 1]
            ratio_rows.append(row_data)
            # Store 5yr avg in kv for easy access
            if val is not None:
                kv[label] = val
            continue

        # Regular key-value pairs
        if val is not None:
            kv[label] = val
    
    # ── Extract structured metrics ──────────────────────────────────────────
    price = parse_dollar(kv.get("Price (As of 04 Mar '26)") or kv.get('Previous Close'))
    # Try B first, then M
    mkt_cap_b_str = kv.get('Market Cap (B)*')
    mkt_cap_m_str = kv.get('Market Cap (M)*')
    if mkt_cap_b_str and str(mkt_cap_b_str).strip() not in ('', '—'):
        mkt_cap_val = parse_dollar(mkt_cap_b_str)
    elif mkt_cap_m_str:
        mkt_cap_val = parse_dollar(mkt_cap_m_str)
        if mkt_cap_val:
            mkt_cap_val = mkt_cap_val / 1000  # Convert M to B
    else:
        mkt_cap_val = None
    
    ev_str = kv.get('Enterprise Value (B)*')
    ev_val = parse_dollar(ev_str)
    
    rating_label, rating_score = parse_rating(kv.get('Avg Rating'))
    target_price = parse_dollar(kv.get('Target Price'))
    broker_count = kv.get('Broker Contributors')
    
    pe_ltm = parse_dollar(kv.get('P/E (LTM)*'))
    ps_ltm = parse_dollar(kv.get('P/S (LTM)*'))
    ev_sales_ltm = parse_dollar(kv.get('EV/Sales (LTM)*'))
    ev_ebitda_ltm = parse_dollar(kv.get('EV/EBITDA (LTM)*'))
    wacc = parse_pct(kv.get('WACC (%)*'))
    beta = parse_dollar(kv.get('Beta (3Y Adj.)*'))
    pct_52w_high = parse_pct(kv.get('% of 52 Week High'))
    
    eps_consensus = parse_dollar(kv.get('EPS Consensus'))
    rev_consensus_str = kv.get('Revenue Consensus (B)') or kv.get('Revenue Consensus (M)') or ''
    rev_consensus = parse_dollar(rev_consensus_str)
    if rev_consensus and 'M' in str(kv.get('Revenue Consensus (M)', '')):
        rev_consensus = rev_consensus / 1000  # normalize to B
    
    next_earnings = kv.get('Next Earnings')
    div_yield = parse_pct(kv.get('Dividend Yield (%)*'))
    annual_div = parse_dollar(kv.get('Annual Dividend*'))
    week52_range = kv.get('52 Week Range')
    volume = kv.get('Volume*')
    avg_vol = kv.get('30D Avg Daily Vol')
    short_interest = kv.get('Short Interest')
    
    # Profitability metrics from financial data
    gross_margin = parse_pct(kv.get('Gross Margin (%)'))
    op_margin = parse_pct(kv.get('Operating Margin (%)'))
    net_margin = parse_pct(kv.get('Net Margin (%)'))
    fcf_margin = parse_pct(kv.get('Free Cash Flow Margin (%)'))
    roe = parse_pct(kv.get('ROE (%)'))
    roa = parse_pct(kv.get('ROA (%)'))
    debt_equity = parse_dollar(kv.get('TDebt%TEQ'))
    current_ratio = parse_dollar(kv.get('Current Ratio'))
    interest_coverage = parse_dollar(kv.get('Interest Coverage'))
    revenue_growth_5yr = parse_dollar(kv.get('Revenue'))  # 5yr CAGR
    pb_ratio = parse_dollar(kv.get('P/BV') or kv.get('P/BV*'))
    pfcf = parse_dollar(kv.get('P/FCF'))
    
    # Revenue segment data
    revenue_fy = kv.get("Revenue as of Sep '25 (FY End)") or kv.get("Revenue as of Jan '26 (FY End)") or kv.get("Revenue as of Jun '25 (FY End)") or kv.get("Revenue as of Dec '25 (FY End)")
    
    # ── Build structured snapshot ───────────────────────────────────────────
    snapshot = {
        'ticker': ticker,
        'name': COMPANY_NAMES.get(ticker, ''),
        'sector': SECTORS.get(ticker, ''),
        'industry': INDUSTRIES.get(ticker, ''),
        'dataSource': 'factset',
        'snapshotDate': '2026-03-05',
        'priceDate': '2026-03-04',
        
        # Trading
        'price': price,
        'week52Range': week52_range,
        'pctOf52wHigh': pct_52w_high,
        'volume': volume,
        'avgVolume30d': avg_vol,
        'shortInterest': short_interest,
        
        # Valuation
        'marketCapB': round(mkt_cap_val, 2) if mkt_cap_val else None,
        'evB': round(ev_val, 2) if ev_val else None,
        'peLTM': pe_ltm,
        'psLTM': ps_ltm,
        'evSalesLTM': ev_sales_ltm,
        'evEbitdaLTM': ev_ebitda_ltm,
        'pbRatio': pb_ratio,
        'pFCF': pfcf,
        'wacc': wacc,
        'beta': beta,
        
        # Estimates / Consensus
        'epsConsensus': eps_consensus,
        'revenueConsensusB': round(rev_consensus, 2) if rev_consensus else None,
        'nextEarnings': next_earnings,
        'analystRating': rating_label,
        'analystRatingScore': rating_score,
        'targetPrice': target_price,
        'brokerContributors': broker_count,
        
        # Dividends
        'annualDividend': annual_div,
        'dividendYield': div_yield,
        
        # Performance
        'pricePerformance': {k: v for k, v in perf.items() if v is not None},
        
        # Profitability
        'grossMargin': gross_margin,
        'operatingMargin': op_margin,
        'netMargin': net_margin,
        'fcfMargin': fcf_margin,
        'roe': roe,
        'roa': roa,
        
        # Solvency
        'debtEquity': debt_equity,
        'currentRatio': current_ratio,
        'interestCoverage': interest_coverage,
        
        # Growth
        'revenueGrowth5yrCAGR': revenue_growth_5yr,
        
        # Time series
        'valuationTimeseries': valuation_rows[:15],
        'financialSummary': financial_rows[:20],
        'ratioTimeseries': ratio_rows[:25],

        # Revenue FY total
        'revenueFYTotal': revenue_fy,
    }

    return snapshot


def create_factset_table(conn):
    """Create the factset_snapshots table if not exists."""
    conn.execute('''
        CREATE TABLE IF NOT EXISTS factset_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            snapshot_date TEXT NOT NULL,
            data_json TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, snapshot_date)
        )
    ''')
    conn.commit()


def ingest_all():
    """Parse all Excel files and ingest into bloomberg_archive.db."""
    conn = sqlite3.connect(DB_PATH)
    create_factset_table(conn)
    
    results = []
    
    for ticker, filepath in EXCEL_FILES.items():
        if not os.path.exists(filepath):
            print(f"  ✗ {ticker}: File not found: {filepath}")
            continue
        
        try:
            snapshot = parse_excel(ticker, filepath)
            data_json = json.dumps(snapshot, ensure_ascii=False, default=str)
            
            # Upsert into factset_snapshots
            conn.execute('''
                INSERT OR REPLACE INTO factset_snapshots (ticker, snapshot_date, data_json)
                VALUES (?, ?, ?)
            ''', (ticker, snapshot['snapshotDate'], data_json))
            
            # Also archive to bbg_reference_archive for backward compatibility
            conn.execute('''
                INSERT INTO bbg_reference_archive (ticker, data_type, fields_json)
                VALUES (?, ?, ?)
            ''', (ticker, 'factset_excel_snapshot', data_json))
            
            conn.commit()
            
            # Summary
            print(f"  ✓ {ticker}: Price=${snapshot['price']}, MktCap=${snapshot['marketCapB']}B, "
                  f"P/E={snapshot['peLTM']}, EV/EBITDA={snapshot['evEbitdaLTM']}, "
                  f"Beta={snapshot['beta']}, Target=${snapshot['targetPrice']}")
            results.append(snapshot)
            
        except Exception as e:
            print(f"  ✗ {ticker}: Parse error: {e}")
            import traceback
            traceback.print_exc()
    
    conn.close()
    return results


if __name__ == '__main__':
    print("═" * 60)
    print("  QuantAlpha — FactSet Excel Ingestion")
    print("═" * 60)
    print(f"  DB: {DB_PATH}")
    print(f"  Files: {len(EXCEL_FILES)} snapshots")
    print()
    
    snapshots = ingest_all()
    
    print()
    print(f"  ✓ Ingested {len(snapshots)} FactSet snapshots into bloomberg_archive.db")
    print(f"  ✓ Tables: factset_snapshots + bbg_reference_archive")
    print("═" * 60)
